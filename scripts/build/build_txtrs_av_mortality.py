#!/usr/bin/env python3
"""Build txtrs-av runtime mortality files.

Active employee mortality:
  - SOA Pub-2010(B) Below-Median-Income Teacher, Employee column, by gender
  - Source-direct from prep/common/sources/soa_pub2010_amount_mort_rates.xlsx
  - The AV's "male set forward 2 years" is applied at runtime via the
    plan_config option male_mp_forward_shift, NOT in this file

Healthy retiree mortality:
  - The valuation names the 2021 TRS of Texas Healthy Pensioner Mortality
    Tables, which are not publicly available
  - Used here: the txtrs-av prototype TX-custom approximation built by
    scripts/build/estimate_txtrs_av_retiree_mortality.py (registry method
    mortality-checkpoint-spline-estimation-v1) and stored at
    prep/txtrs-av/reference_tables/estimated_retiree_mortality_base_2021.csv
  - Backprojected from 2021 to 2010 using UMP-2021 ultimate rates so the
    runtime, which assumes base_year=2010, projects forward to the prototype's
    fitted 2021 values

Improvement scale:
  - UMP-2021 ultimate rates ("immediate convergence" interpretation of the
    AV's Scale UMP 2021 wording)
  - Read directly from prep/common/sources/soa_mp2021_rates.xlsx, last column
    (2037+ ultimate)
  - Stored flat across years 1951-2037 in long format

This script does not reuse plans/txtrs/data/mortality/ files.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd


PROJECT_ROOT = Path(__file__).resolve().parents[2]

PUB_PATH = PROJECT_ROOT / "prep" / "common" / "sources" / "soa_pub2010_amount_mort_rates.xlsx"
MP_PATH = PROJECT_ROOT / "prep" / "common" / "sources" / "soa_mp2021_rates.xlsx"
PROTO_PATH = (
    PROJECT_ROOT
    / "prep"
    / "txtrs-av"
    / "reference_tables"
    / "estimated_retiree_mortality_base_2021.csv"
)
OUT_BASE = PROJECT_ROOT / "plans" / "txtrs-av" / "data" / "mortality" / "base_rates.csv"
OUT_IMP = PROJECT_ROOT / "plans" / "txtrs-av" / "data" / "mortality" / "improvement_scale.csv"

TABLE_NAME = "txtrs_av_av_first"
RUNTIME_BASE_YEAR = 2010
PROTO_BASE_YEAR = 2021
MIN_AGE = 18
MAX_AGE = 120

# SOA MP-2021 publishes 1951..2037+ (the ultimate column is 2037+).
SCALE_YEAR_MIN = 1951
SCALE_YEAR_MAX = 2037

# Layout of the PubT-2010(B) "Teachers - Below-Median Income" sheet:
#   row 4 (zero-indexed) is the column-name header row with
#   col 1: Age, col 3: Female Employee, col 4: Female Healthy Retiree,
#   col 7: Male Employee,  col 8: Male Healthy Retiree.
PUB_SHEET = "PubT-2010(B)"
PUB_DATA_FIRST_ROW = 5
PUB_COL_AGE = 1
PUB_COL_FEMALE_EMPLOYEE = 3
PUB_COL_MALE_EMPLOYEE = 7


def _read_pub_2010b_employee() -> pd.DataFrame:
    """Read Pub-2010(B) Employee qx by age and gender for ages 18-MAX_AGE.

    The published Employee column covers ages 18 through 80; older rows are
    blank in the source workbook.
    """
    wb = openpyxl.load_workbook(PUB_PATH, data_only=True, read_only=True)
    ws = wb[PUB_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    out: list[dict] = []
    for r in rows[PUB_DATA_FIRST_ROW:]:
        age = r[PUB_COL_AGE]
        if not isinstance(age, int):
            continue
        if age < MIN_AGE or age > MAX_AGE:
            continue
        for gender, col in (
            ("female", PUB_COL_FEMALE_EMPLOYEE),
            ("male", PUB_COL_MALE_EMPLOYEE),
        ):
            qx = r[col]
            if qx in ("", None):
                continue
            out.append({"age": age, "gender": gender, "qx": float(qx)})

    return pd.DataFrame(out)


def _read_mp2021_ultimate() -> pd.DataFrame:
    """Read UMP-2021 ultimate (2037+) annual mortality improvement rate by
    age and gender. UMP-2021 ultimate rates are unisex by age, but we read
    each gender sheet separately so the file structure dictates the values
    rather than assuming."""
    wb = openpyxl.load_workbook(MP_PATH, data_only=True, read_only=True)
    out: list[dict] = []
    for sheet, gender in (("Female", "female"), ("Male", "male")):
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        headers = list(rows[1])
        last_col = max(i for i, h in enumerate(headers) if h is not None)
        for r in rows[2:]:
            age_cell = r[0]
            if isinstance(age_cell, str) and age_cell.startswith("≤"):
                age = int(age_cell.split()[-1])
            elif isinstance(age_cell, int):
                age = age_cell
            else:
                continue
            ultimate = r[last_col]
            if ultimate is None:
                continue
            out.append({"age": age, "gender": gender, "improvement": float(ultimate)})
    wb.close()
    return (
        pd.DataFrame(out)
        .drop_duplicates(subset=["age", "gender"])
        .sort_values(["gender", "age"])
        .reset_index(drop=True)
    )


def build_retiree_qx_2010(prototype: pd.DataFrame, ultimate: pd.DataFrame) -> pd.DataFrame:
    """Backproject prototype 2021 retiree qx to 2010.

    qx_2010 = qx_2021 / (1 - improvement)^(PROTO_BASE_YEAR - RUNTIME_BASE_YEAR)

    The runtime mortality builder assumes base_year=2010. Storing the
    prototype's 2021 fit directly would cause the runtime to over-improve to
    a 2032 effective table. Backprojecting cancels that out so the runtime,
    when it walks 2010 -> 2021 with the same UMP-2021 ultimate rates,
    reproduces the prototype's published-checkpoint-anchored 2021 curve.
    """
    ult = ultimate.set_index(["age", "gender"])["improvement"]
    n_years = PROTO_BASE_YEAR - RUNTIME_BASE_YEAR

    rows: list[dict] = []
    for _, r in prototype.iterrows():
        age = int(r["age"])
        for gender, qx_col in (
            ("female", "estimated_2021_female_qx"),
            ("male", "estimated_2021_male_qx"),
        ):
            qx_2021 = float(r[qx_col])
            imp = float(ult.get((age, gender), 0.0))
            denom = (1.0 - imp) ** n_years
            qx_2010 = qx_2021 / denom if denom > 0 else qx_2021
            rows.append({"age": age, "gender": gender, "qx": min(qx_2010, 1.0)})

    return pd.DataFrame(rows)


def build_base_rates(employee_qx: pd.DataFrame, retiree_qx: pd.DataFrame) -> pd.DataFrame:
    """Assemble base_rates.csv long-format with columns
    age, gender, member_type, table, qx for ages 18-120.

    Where one member_type is missing for an (age, gender), fill from the
    other; this matches the runtime _read_base_mort_csv NaN-fill rule.
    """
    employee = employee_qx.assign(member_type="employee")
    retiree = retiree_qx.assign(member_type="retiree")
    combined = pd.concat([employee, retiree], ignore_index=True)

    full = pd.MultiIndex.from_product(
        [range(MIN_AGE, MAX_AGE + 1), ["female", "male"], ["employee", "retiree"]],
        names=["age", "gender", "member_type"],
    ).to_frame(index=False)
    merged = full.merge(combined, on=["age", "gender", "member_type"], how="left")

    pivot = merged.pivot_table(
        index=["age", "gender"], columns="member_type", values="qx", aggfunc="first"
    ).reset_index()
    pivot["employee"] = pivot["employee"].fillna(pivot["retiree"])
    pivot["retiree"] = pivot["retiree"].fillna(pivot["employee"])

    melted = pivot.melt(
        id_vars=["age", "gender"],
        value_vars=["employee", "retiree"],
        var_name="member_type",
        value_name="qx",
    )
    melted["table"] = TABLE_NAME
    return (
        melted[["age", "gender", "member_type", "table", "qx"]]
        .sort_values(["member_type", "gender", "age"])
        .reset_index(drop=True)
    )


def build_improvement_scale(ultimate: pd.DataFrame) -> pd.DataFrame:
    """Replicate UMP-2021 ultimate rates flat across years 1951-2037."""
    rows: list[dict] = []
    for _, r in ultimate.iterrows():
        age = int(r["age"])
        gender = r["gender"]
        imp = float(r["improvement"])
        for year in range(SCALE_YEAR_MIN, SCALE_YEAR_MAX + 1):
            rows.append({"age": age, "year": year, "gender": gender, "improvement": imp})
    return (
        pd.DataFrame(rows)
        .sort_values(["gender", "age", "year"])
        .reset_index(drop=True)
    )


def _write_csv(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False)
    print(f"Wrote {path}")


def main() -> None:
    employee_qx = _read_pub_2010b_employee()
    ultimate = _read_mp2021_ultimate()
    prototype = pd.read_csv(PROTO_PATH)
    retiree_qx = build_retiree_qx_2010(prototype, ultimate)

    base_rates = build_base_rates(employee_qx, retiree_qx)
    improvement = build_improvement_scale(ultimate)

    _write_csv(base_rates, OUT_BASE)
    _write_csv(improvement, OUT_IMP)


if __name__ == "__main__":
    main()
