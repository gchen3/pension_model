"""
Truth tables: plan-level aggregates for visual R-vs-Python comparison.

A truth table is a single-sheet DataFrame summarizing one pension plan's key
outputs year-by-year at the plan-wide level. It does NOT diagnose why things
differ; it tells you quickly WHETHER anything differs materially from the R
baseline.

Two views of the same data live side by side in an Excel workbook:
- `frs_R`, `txtrs_R`: frozen values extracted from the R model's baseline
  CSVs, written once by scripts/build_r_truth_tables.py and never re-generated
- `frs_Py`, `txtrs_Py`: overwritten by each Python pipeline run (via the CLI),
  so flipping between tabs gives a direct visual diff

Column layout — organized so the MVA balance identity is visible:

  BALANCES
    mva_boy              — market value of DB assets at beginning of year

  INFLOWS (into DB fund)
    er_db_cont           — employer DB contributions (normal cost + amortization;
                           includes DROP employer contributions for FRS)
    ee_cont              — employee contributions
    invest_income        — actual investment income earned on MVA

  OUTFLOWS (from DB fund)
    benefits             — benefit payments (includes DROP benefit payments for FRS)
    refunds              — refund payments to terminated members
    admin_exp            — administrative expenses

  RESULT
    mva_eoy              — end-of-year MVA; equals next row's mva_boy.
                           For projected years (not the init row):
                           mva_eoy = mva_boy + er_db_cont + ee_cont + invest_income
                                     - benefits - refunds - admin_exp

  OTHER BALANCES
    aal_boy              — actuarial accrued liability at BOY
    ava_boy              — actuarial value of assets at BOY
    fr_mva_boy           — funded ratio = mva_boy / aal_boy
    fr_ava_boy           — funded ratio = ava_boy / aal_boy

  DEMOGRAPHICS
    n_active_boy         — active member headcount at BOY
    payroll              — total payroll during fiscal year
    er_cont_total        — total employer contributions (DB + DC), for reference

Note: DC employer contributions (er_dc_cont) are not in the MVA balance because
DC money is paid directly to member accounts and never flows through the DB fund.
For FRS, DROP flows are included in the aggregate er_db_cont and benefits columns.

MVA balance identity (holds to float precision for all rows except the last):
    mva_eoy = mva_boy + er_db_cont + ee_cont + invest_income
              - benefits - refunds - admin_exp
    mva_eoy should equal next row's mva_boy.

Missing values are reported as pandas NA and will round-trip as empty cells in
both CSV and Excel.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict

import pandas as pd


# Canonical column order for every truth table (R and Python, FRS and TRS).
TRUTH_TABLE_COLUMNS = [
    "plan",
    # --- Balances ---
    "year",
    "mva_boy",
    # --- Inflows (to DB fund) ---
    "er_db_cont",
    "ee_cont",
    "invest_income",
    # --- Outflows (from DB fund) ---
    "benefits",
    "refunds",
    "admin_exp",
    # --- Result: mva_eoy = mva_boy + inflows - outflows ---
    "mva_eoy",
    # --- Other balances ---
    "aal_boy",
    "ava_boy",
    "fr_mva_boy",
    "fr_ava_boy",
    # --- Demographics ---
    "n_active_boy",
    "payroll",
    "er_cont_total",
]

import numpy as np


def _actual_invest_income(mva, net_cf):
    """Compute actual investment income so the MVA balance identity holds.

    The truth table identity is:
      mva_boy[i+1] = mva_boy[i] + er_db_cont[i] + ee_cont[i]
                     + invest_income[i] - benefits[i] - refunds[i] - admin[i]

    Since net_cf[i] = er_db_cont[i] + ee_cont[i] - benefits[i] - refunds[i] - admin[i],
    this reduces to: invest_income[i] = mva_boy[i+1] - mva_boy[i] - net_cf[i]

    The last row gets 0 (no mva_boy[n] to compare against).
    """
    n = len(mva)
    result = np.zeros(n)
    for i in range(n - 1):
        result[i] = mva[i + 1] - mva[i] - net_cf[i]
    return result


# ---------------------------------------------------------------------------
# R-side builders — read from R output CSVs, produce a truth table
# ---------------------------------------------------------------------------

def build_r_truth_table_frs(baseline_dir: Path) -> pd.DataFrame:
    """Build the frozen FRS R truth table from the R model's output CSVs.

    Sources:
      - `frs_funding.csv`: plan-wide aggregate of 7 classes + DROP (payroll,
        AAL, MVA, AVA, contributions, benefits, investment income, funded ratios)
      - `{class}_liability.csv` × 7: per-class total_n_active, summed for
        plan-wide n_active

    Not reported (NA) because R does not emit them:
      - n_retired_boy, n_inactive_boy
    """
    f = pd.read_csv(baseline_dir / "frs_funding.csv")

    # Sum active headcount across all 7 FRS classes (per-year)
    frs_classes = ["regular", "special", "admin", "eco", "eso",
                   "judges", "senior_management"]
    n_active = None
    for cn in frs_classes:
        liab = pd.read_csv(baseline_dir / f"{cn}_liability.csv")
        col = liab["total_n_active"].values
        n_active = col if n_active is None else n_active + col

    net_cf = f["net_cf_legacy"].values + f["net_cf_new"].values
    mva = f["total_mva"].values
    invest_income = _actual_invest_income(mva, net_cf)
    ee = f["ee_nc_cont_legacy"].values + f["ee_nc_cont_new"].values
    er_db = f["total_er_db_cont"].values
    benefits = f["total_ben_payment"].values
    refunds = f["total_refund"].values
    admin = f["admin_exp_legacy"].values + f["admin_exp_new"].values

    df = pd.DataFrame({
        "plan": "frs",
        "year": f["year"].astype(int).values,
        "mva_boy": mva,
        "er_db_cont": er_db,
        "ee_cont": ee,
        "invest_income": invest_income,
        "benefits": benefits,
        "refunds": refunds,
        "admin_exp": admin,
        "mva_eoy": mva + net_cf + invest_income,
        "aal_boy": f["total_aal"].values,
        "ava_boy": f["total_ava"].values,
        "fr_mva_boy": f["fr_mva"].values,
        "fr_ava_boy": f["fr_ava"].values,
        "n_active_boy": n_active,
        "payroll": f["total_payroll"].values,
        "er_cont_total": f["total_er_cont"].values,
    })
    return df[TRUTH_TABLE_COLUMNS]


def build_r_truth_table_txtrs(trs_r_dir: Path) -> pd.DataFrame:
    """Build the frozen TRS R truth table from the R model's output CSVs.

    Sources:
      - `baseline_fresh.csv`: liability output (n.active)
      - `funding_fresh.csv`: funding output (payroll, AAL, MVA, AVA,
        contributions, benefits, investment income, funded ratios)

    Not reported (NA) because R does not emit them:
      - n_retired_boy, n_inactive_boy
    """
    liab = pd.read_csv(trs_r_dir / "baseline_fresh.csv")
    f = pd.read_csv(trs_r_dir / "funding_fresh.csv")

    net_cf = f["net_cf_legacy"].fillna(0).values + f["net_cf_new"].fillna(0).values
    mva = f["MVA"].values
    invest_income = _actual_invest_income(mva, net_cf)
    er_db = f["er_cont"].values  # TRS has no DC plan
    ee = f["ee_nc_cont_legacy"].fillna(0).values + f["ee_nc_cont_new"].fillna(0).values
    benefits = f["ben_payment_legacy"].fillna(0).values + f["ben_payment_new"].fillna(0).values
    refunds = f["refund_legacy"].fillna(0).values + f["refund_new"].fillna(0).values
    admin = f["admin_exp_legacy"].fillna(0).values + f["admin_exp_new"].fillna(0).values

    df = pd.DataFrame({
        "plan": "txtrs",
        "year": f["fy"].astype(int).values,
        "mva_boy": mva,
        "er_db_cont": er_db,
        "ee_cont": ee,
        "invest_income": invest_income,
        "benefits": benefits,
        "refunds": refunds,
        "admin_exp": admin,
        "mva_eoy": mva + net_cf + invest_income,
        "aal_boy": f["AAL"].values,
        "ava_boy": f["AVA"].values,
        "fr_mva_boy": f["FR_MVA"].values,
        "fr_ava_boy": f["FR_AVA"].values,
        "n_active_boy": liab["n.active"].values,
        "payroll": f["payroll"].values,
        "er_cont_total": er_db,  # same as er_db for TRS (no DC)
    })
    return df[TRUTH_TABLE_COLUMNS]


# ---------------------------------------------------------------------------
# Python-side builder — take live pipeline output, produce a truth table
# ---------------------------------------------------------------------------

def build_python_truth_table(
    plan_name: str,
    liability: Dict[str, pd.DataFrame],
    funding,
    constants,
) -> pd.DataFrame:
    """Build a Python-side truth table from live pipeline output.

    Args:
        plan_name: "frs" or "txtrs".
        liability: dict of class_name -> liability DataFrame (per-class).
        funding: the funding object — for FRS this is a dict with a "frs"
            key aggregated across all classes; for TRS it's a single
            DataFrame for the "all" class.
        constants: PlanConfig for the plan.

    Metrics the Python pipeline does not (yet) compute are returned as NA.
    """
    fmt = constants.raw.get("truth_table_format", plan_name)
    if fmt == "frs":
        return _build_python_truth_table_frs(liability, funding, constants)
    elif fmt == "txtrs":
        return _build_python_truth_table_txtrs(liability, funding, constants)
    else:
        raise ValueError(f"unknown truth_table_format: {fmt!r}")


def _build_python_truth_table_frs(liability, funding, constants) -> pd.DataFrame:
    """FRS-style: funding is a dict containing plan aggregate."""
    f = funding[constants.plan_name]

    # Sum total_n_active across all classes
    classes = list(constants.classes)
    n_active = None
    for cn in classes:
        col = liability[cn]["total_n_active"].values
        n_active = col if n_active is None else n_active + col

    net_cf = f["net_cf_legacy"].values + f["net_cf_new"].values
    mva = f["total_mva"].values
    invest_income = _actual_invest_income(mva, net_cf)
    ee = f["ee_nc_cont_legacy"].values + f["ee_nc_cont_new"].values
    er_db = f["total_er_db_cont"].values
    benefits = f["total_ben_payment"].values
    refunds = f["total_refund"].values
    admin = f["admin_exp_legacy"].values + f["admin_exp_new"].values

    df = pd.DataFrame({
        "plan": "frs",
        "year": f["year"].astype(int).values,
        "mva_boy": mva,
        "er_db_cont": er_db,
        "ee_cont": ee,
        "invest_income": invest_income,
        "benefits": benefits,
        "refunds": refunds,
        "admin_exp": admin,
        "mva_eoy": mva + net_cf + invest_income,
        "aal_boy": f["total_aal"].values,
        "ava_boy": f["total_ava"].values,
        "fr_mva_boy": f["fr_mva"].values,
        "fr_ava_boy": f["fr_ava"].values,
        "n_active_boy": n_active,
        "payroll": f["total_payroll"].values,
        "er_cont_total": f["total_er_cont"].values,
    })
    return df[TRUTH_TABLE_COLUMNS]


def _build_python_truth_table_txtrs(liability, funding, _constants) -> pd.DataFrame:
    """TRS: funding is a single DataFrame; liability['all'] is the per-class frame."""
    f = funding
    liab = liability["all"]

    def col(df, *options):
        for o in options:
            if o in df.columns:
                return df[o].values
        return None

    def _sum(a, b):
        if a is not None and b is not None:
            return a + b
        return a if a is not None else b

    year = col(f, "year")
    mva = col(f, "total_mva")
    er_db = col(f, "total_er_cont")  # TRS has no DC
    ee = _sum(col(f, "ee_nc_cont_legacy"), col(f, "ee_nc_cont_new"))
    benefits = _sum(col(f, "ben_payment_legacy"), col(f, "ben_payment_new"))
    if benefits is None:
        benefits = col(f, "total_ben_payment")
    refunds = _sum(col(f, "refund_legacy"), col(f, "refund_new"))
    if refunds is None:
        refunds = col(f, "total_refund")
    admin = _sum(col(f, "admin_exp_legacy"), col(f, "admin_exp_new"))
    net_cf = _sum(col(f, "net_cf_legacy"), col(f, "net_cf_new"))

    invest_income = _actual_invest_income(mva, net_cf) if mva is not None and net_cf is not None else None

    n = len(year) if year is not None else len(liab)
    z = np.zeros(n)

    _or_z = lambda x: x if x is not None else z

    df = pd.DataFrame({
        "plan": "txtrs",
        "year": pd.Series(year).astype(int).values,
        "mva_boy": mva,
        "er_db_cont": _or_z(er_db),
        "ee_cont": _or_z(ee),
        "invest_income": _or_z(invest_income),
        "benefits": _or_z(benefits),
        "refunds": _or_z(refunds),
        "admin_exp": _or_z(admin),
        "mva_eoy": _or_z(mva) + _or_z(net_cf) + _or_z(invest_income),
        "aal_boy": col(f, "total_aal"),
        "ava_boy": col(f, "total_ava"),
        "fr_mva_boy": col(f, "fr_mva"),
        "fr_ava_boy": col(f, "fr_ava"),
        "n_active_boy": col(liab, "total_n_active"),
        "payroll": col(f, "total_payroll"),
        "er_cont_total": _or_z(er_db),  # same as er_db for TRS
    })
    return df[TRUTH_TABLE_COLUMNS]


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------

def format_truth_table_for_log(df: pd.DataFrame, max_rows: int = 31) -> str:
    """Render a truth table as a human-readable text block for stdout/logs.

    Dollar values are shown in millions with comma separators; counts are
    shown as integers; funded ratios as percentages. NA shows as a dash.
    """
    df = df.head(max_rows).copy()

    def _fmt_dollars(v):
        if pd.isna(v):
            return "    —"
        return f"{v / 1e6:>10,.0f}"

    def _fmt_count(v):
        if pd.isna(v):
            return "       —"
        return f"{v:>8,.0f}"

    def _fmt_pct(v):
        if pd.isna(v):
            return "    —"
        return f"{v * 100:>6.2f}%"

    header = (
        f"  {'year':>4s} {'mva_boy':>10s} {'er_db':>10s} {'ee':>10s} "
        f"{'invest':>10s} {'benefits':>10s} {'refunds':>10s} {'admin':>10s} "
        f"{'mva_eoy':>10s} {'aal_boy':>10s} {'fr_mva':>7s} {'fr_ava':>7s} "
        f"{'active':>8s} {'payroll':>10s}"
    )
    sep = "  " + "-" * (len(header) - 2)
    lines = [header, sep]

    for _, row in df.iterrows():
        line = (
            f"  {int(row['year']):>4d} "
            f"{_fmt_dollars(row['mva_boy'])} "
            f"{_fmt_dollars(row['er_db_cont'])} "
            f"{_fmt_dollars(row['ee_cont'])} "
            f"{_fmt_dollars(row['invest_income'])} "
            f"{_fmt_dollars(row['benefits'])} "
            f"{_fmt_dollars(row['refunds'])} "
            f"{_fmt_dollars(row['admin_exp'])} "
            f"{_fmt_dollars(row['mva_eoy'])} "
            f"{_fmt_dollars(row['aal_boy'])} "
            f"{_fmt_pct(row['fr_mva_boy'])} "
            f"{_fmt_pct(row['fr_ava_boy'])} "
            f"{_fmt_count(row['n_active_boy'])} "
            f"{_fmt_dollars(row['payroll'])}"
        )
        lines.append(line)

    lines.append("")
    lines.append("  (dollar amounts in millions; funded ratios as percentages)")
    lines.append("  mva_eoy = mva_boy + er_db + ee + invest - benefits - refunds - admin")
    return "\n".join(lines)


def _freeze_panes(ws, df):
    """Freeze below header row and right of label columns (plan, year)."""
    label_cols = sum(1 for c in df.columns if c in ("plan", "year"))
    # +1 for the Excel 1-based index, +1 to freeze right of the last label col
    col_letter = chr(ord("A") + label_cols)
    ws.freeze_panes = f"{col_letter}2"


def upsert_sheet_to_excel(df: pd.DataFrame, xlsx_path: Path, sheet_name: str) -> None:
    """Write `df` to `sheet_name` in `xlsx_path`, creating the file if needed
    and preserving any other sheets that already exist.

    This is the function the CLI uses to write a Python sheet without
    clobbering the frozen R sheets in the same workbook.
    """
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    if xlsx_path.exists():
        with pd.ExcelWriter(
            xlsx_path, engine="openpyxl",
            mode="a", if_sheet_exists="replace",
        ) as w:
            df.to_excel(w, sheet_name=sheet_name, index=False)
            _freeze_panes(w.sheets[sheet_name], df)
    else:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
            df.to_excel(w, sheet_name=sheet_name, index=False)
            _freeze_panes(w.sheets[sheet_name], df)


# ---------------------------------------------------------------------------
# Diff sheet — live formulas that recompute when the Py sheet is overwritten
# ---------------------------------------------------------------------------

# Metrics in the diff sheet. plan/year are label columns, not metrics.
_DIFF_METRICS = [
    ("n_active_boy", "active"),
    ("n_retired_boy", "retired"),
    ("n_inactive_boy", "inactive"),
    ("payroll_fy", "payroll"),
    ("benefits_fy", "benefits"),
    ("aal_boy", "aal"),
    ("er_cont_fy", "er_cont"),
    ("ee_cont_fy", "ee_cont"),
    ("mva_boy", "mva"),
    ("invest_income_fy", "inv_income"),
    ("ava_boy", "ava"),
    ("fr_mva_boy", "fr_mva"),
    ("fr_ava_boy", "fr_ava"),
]


def write_diff_sheet_with_formulas(
    xlsx_path: Path,
    diff_sheet_name: str,
    r_sheet_name: str,
    py_sheet_name: str,
    n_rows: int,
) -> None:
    """Write a diff sheet whose cells are live formulas referencing the R
    and Py sheets, so every time the Py sheet is overwritten by a pipeline
    run the diffs update automatically in Excel.

    Layout (per year row, 41 columns total):
      A: plan          (=r_sheet!A2)   — pulled from R sheet by formula
      B: year          (=r_sheet!B2)
      C: active_R      (=r_sheet!C2)
      D: active_Py     (=py_sheet!C2)
      E: active_diff   (=IFERROR(D2-C2, ""))
      F: retired_R     ...
      ...

    Diff is an ABSOLUTE difference (Py - R), not a percentage.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
    from openpyxl.formatting.rule import CellIsRule

    wb = load_workbook(xlsx_path)

    # Drop existing diff sheet if present — we rewrite from scratch
    if diff_sheet_name in wb.sheetnames:
        del wb[diff_sheet_name]
    ws = wb.create_sheet(diff_sheet_name)

    # --- Header row 1 (group labels spanning 3 cols) + row 2 (R/Py/diff) ---
    # Row 1: "plan" | "year" | metric label across 3 cells | ...
    # Row 2: blank  | blank  | R | Py | diff | R | Py | diff | ...
    ws.cell(row=1, column=1, value="").font = Font(bold=True)
    ws.cell(row=1, column=2, value="").font = Font(bold=True)
    ws.cell(row=2, column=1, value="plan").font = Font(bold=True)
    ws.cell(row=2, column=2, value="year").font = Font(bold=True)

    for i, (_src_col, label) in enumerate(_DIFF_METRICS):
        c_r = 3 + i * 3        # R column
        c_py = c_r + 1         # Py column
        c_diff = c_r + 2       # diff column
        # Merge three cells in header row 1 for the metric label
        ws.cell(row=1, column=c_r, value=label).font = Font(bold=True)
        ws.cell(row=1, column=c_r).alignment = Alignment(horizontal="center")
        ws.merge_cells(
            start_row=1, start_column=c_r, end_row=1, end_column=c_diff
        )
        # Row 2 sub-labels
        ws.cell(row=2, column=c_r, value="R").font = Font(bold=True, italic=True)
        ws.cell(row=2, column=c_py, value="Py").font = Font(bold=True, italic=True)
        ws.cell(row=2, column=c_diff, value="diff").font = Font(bold=True, italic=True)

    # --- Data rows: formulas referencing r_sheet_name and py_sheet_name ---
    # Source sheets have header in row 1 and data starting at row 2.
    # Column mapping in the source sheets (same for R and Py):
    #   A=plan, B=year, C=n_active_boy, D=n_retired_boy, E=n_inactive_boy,
    #   F=payroll_fy, G=benefits_fy, H=aal_boy, I=er_cont_fy, J=ee_cont_fy,
    #   K=mva_boy, L=invest_income_fy, M=ava_boy, N=fr_mva_boy, O=fr_ava_boy
    #
    # Source col C is metric index 0, D is metric index 1, etc.
    for row_idx in range(n_rows):
        src_row = row_idx + 2       # source data starts at row 2
        dst_row = row_idx + 3       # diff data starts at row 3 (after 2 header rows)

        # plan and year — pull from R sheet by formula
        ws.cell(row=dst_row, column=1,
                value=f"='{r_sheet_name}'!A{src_row}")
        ws.cell(row=dst_row, column=2,
                value=f"='{r_sheet_name}'!B{src_row}")

        for i, _ in enumerate(_DIFF_METRICS):
            src_col_letter = get_column_letter(3 + i)   # C, D, E, ...
            c_r = 3 + i * 3
            c_py = c_r + 1
            c_diff = c_r + 2

            ws.cell(row=dst_row, column=c_r,
                    value=f"='{r_sheet_name}'!{src_col_letter}{src_row}")
            ws.cell(row=dst_row, column=c_py,
                    value=f"='{py_sheet_name}'!{src_col_letter}{src_row}")
            # IFERROR wraps NA handling (blank cell in either side -> blank diff)
            ws.cell(
                row=dst_row, column=c_diff,
                value=(
                    f'=IFERROR('
                    f"'{py_sheet_name}'!{src_col_letter}{src_row}"
                    f"-'{r_sheet_name}'!{src_col_letter}{src_row}"
                    f',"")'
                ),
            )

    # --- Number formatting ---
    # Counts (active/retired/inactive, metric idx 0-2): integer with commas
    # Dollars (payroll through ava, metric idx 3-10): integer with commas (raw $)
    # Funded ratios (fr_mva, fr_ava, metric idx 11-12): percentage w/ 4 decimals
    count_metrics = {0, 1, 2}
    ratio_metrics = {11, 12}
    count_fmt = "#,##0;(#,##0)"
    dollar_fmt = '"$"#,##0;("$"#,##0)'
    ratio_fmt = "0.0000%;(0.0000%)"
    diff_count_fmt = "#,##0;(#,##0)"
    diff_dollar_fmt = '"$"#,##0;("$"#,##0)'
    diff_ratio_fmt = "0.000000;(0.000000)"  # absolute ratio diff, 6 decimals

    for i in range(len(_DIFF_METRICS)):
        c_r = 3 + i * 3
        c_py = c_r + 1
        c_diff = c_r + 2
        if i in count_metrics:
            val_fmt, dif_fmt = count_fmt, diff_count_fmt
        elif i in ratio_metrics:
            val_fmt, dif_fmt = ratio_fmt, diff_ratio_fmt
        else:
            val_fmt, dif_fmt = dollar_fmt, diff_dollar_fmt
        for r_row in range(3, 3 + n_rows):
            ws.cell(row=r_row, column=c_r).number_format = val_fmt
            ws.cell(row=r_row, column=c_py).number_format = val_fmt
            ws.cell(row=r_row, column=c_diff).number_format = dif_fmt

    # --- Column widths ---
    # plan, year: narrow. Metric value cols: wide enough for $ figures. diff
    # cols: same width as value cols so differences are easy to eyeball.
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 6
    for i in range(len(_DIFF_METRICS)):
        base = 3 + i * 3
        for offset in range(3):
            col_letter = get_column_letter(base + offset)
            # Wider for dollar columns, narrower for counts/ratios
            if i in count_metrics:
                ws.column_dimensions[col_letter].width = 12
            elif i in ratio_metrics:
                ws.column_dimensions[col_letter].width = 13
            else:
                ws.column_dimensions[col_letter].width = 17

    # Freeze top 2 header rows and first 2 label columns
    ws.freeze_panes = "C3"

    # --- Conditional formatting: highlight diff cells that are materially
    # nonzero. Since the user wants raw diffs (not percents) we can't use a
    # single threshold across metrics — dollars and ratios are different
    # scales. Apply per-column thresholds.
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for i in range(len(_DIFF_METRICS)):
        c_diff = 3 + i * 3 + 2
        col_letter = get_column_letter(c_diff)
        diff_range = f"{col_letter}3:{col_letter}{2 + n_rows}"
        if i in count_metrics:
            # Counts: red if abs diff > 100, yellow if > 1
            red_thresh, yel_thresh = 100, 1
        elif i in ratio_metrics:
            # Ratios: red if abs diff > 0.001 (0.1 pt), yellow if > 0.00001
            red_thresh, yel_thresh = 0.001, 0.00001
        else:
            # Dollars: red if abs diff > $1M, yellow if > $1
            red_thresh, yel_thresh = 1_000_000, 1
        ws.conditional_formatting.add(
            diff_range,
            CellIsRule(operator="greaterThan", formula=[str(red_thresh)], fill=red_fill),
        )
        ws.conditional_formatting.add(
            diff_range,
            CellIsRule(operator="lessThan", formula=[str(-red_thresh)], fill=red_fill),
        )
        ws.conditional_formatting.add(
            diff_range,
            CellIsRule(operator="greaterThan", formula=[str(yel_thresh)], fill=yellow_fill),
        )
        ws.conditional_formatting.add(
            diff_range,
            CellIsRule(operator="lessThan", formula=[str(-yel_thresh)], fill=yellow_fill),
        )

    wb.save(xlsx_path)
